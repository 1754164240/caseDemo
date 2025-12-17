from typing import List, Optional
import asyncio
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.db.session import get_db, SessionLocal
from app.api.deps import get_current_active_user
from app.models.user import User
from app.models.test_case import TestCase
from app.models.test_point import TestPoint
from app.models.requirement import Requirement
from app.models.scenario import Scenario
from app.models.system_config import SystemConfig
from app.schemas.test_case import TestCase as TestCaseSchema, TestCaseCreate, TestCaseUpdate, TestCaseApproval
from app.schemas.common import PaginatedResponse
from app.services.ai_service import get_ai_service
from app.services.websocket_service import manager
from app.services.document_parser import DocumentParser
from app.services.automation_service import automation_service

router = APIRouter()


def _run_async_notification(loop: Optional[asyncio.AbstractEventLoop], coro, description: str):
    """在线程池中运行的任务里安全地发送异步通知"""
    if not loop:
        return
    try:
        future = asyncio.run_coroutine_threadsafe(coro, loop)
        future.result()
    except Exception as notify_error:
        print(f"[WARNING] {description}: {notify_error}")


def generate_test_case_code(db: Session, test_point: TestPoint) -> str:
    """生成测试用例编号"""
    # 获取该测试点下的测试用例数量
    count = db.query(func.count(TestCase.id)).filter(
        TestCase.test_point_id == test_point.id
    ).scalar()

    # 生成编号：测试点编号-序号
    return f"{test_point.code}-{count + 1}"


def generate_test_cases_background(
    test_point_id: int,
    user_id: int,
    loop: Optional[asyncio.AbstractEventLoop] = None
):
    """后台生成测试用例（在线程池运行）"""
    db = SessionLocal()
    try:
        test_point = db.query(TestPoint).filter(TestPoint.id == test_point_id).first()
        if not test_point:
            return
        
        requirement = db.query(Requirement).filter(Requirement.id == test_point.requirement_id).first()
        if not requirement:
            return
        
        # 解析需求文档获取上下文
        context = DocumentParser.parse(requirement.file_path, requirement.file_type.value)
        
        # 准备测试点数据
        test_point_data = {
            'title': test_point.title,
            'description': test_point.description,
            'category': test_point.category,
            'priority': test_point.priority,
            'business_line': test_point.business_line  # 添加业务线字段
        }

        # 使用 AI 生成测试用例
        ai_svc = get_ai_service(db)
        test_cases_data = ai_svc.generate_test_cases(test_point_data, context)
        
        # 保存测试用例
        for tc_data in test_cases_data:
            code = generate_test_case_code(db, test_point)

            # 处理 preconditions: 如果是字典，转换为 JSON 字符串
            preconditions = tc_data.get('preconditions', '')
            if isinstance(preconditions, dict):
                import json
                preconditions = json.dumps(preconditions, ensure_ascii=False)

            test_case = TestCase(
                test_point_id=test_point_id,
                code=code,
                title=tc_data.get('title', ''),
                description=tc_data.get('description', ''),
                preconditions=preconditions,
                test_steps=tc_data.get('test_steps', []),
                expected_result=tc_data.get('expected_result', ''),
                priority=tc_data.get('priority', 'medium'),
                test_type=tc_data.get('test_type', 'functional')
            )
            db.add(test_case)
            db.flush()  # 确保编号被保存

        db.commit()
        
        # 发送通知
        _run_async_notification(
            loop,
            manager.notify_test_case_generated(user_id, test_point_id, len(test_cases_data)),
            "发送测试用例生成通知失败"
        )
        
    except Exception as e:
        db.rollback()
        print(f"[ERROR] 生成测试用例失败: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


@router.get("/", response_model=PaginatedResponse[TestCaseSchema])
def read_test_cases(
    requirement_id: int = None,
    test_point_id: int = None,
    search: str = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """获取测试用例列表"""
    query = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        Requirement.user_id == current_user.id
    )

    # 按需求筛选
    if requirement_id:
        query = query.filter(Requirement.id == requirement_id)

    # 按测试点筛选
    if test_point_id:
        query = query.filter(TestCase.test_point_id == test_point_id)

    # 搜索功能
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            (TestCase.title.ilike(search_pattern)) |
            (TestCase.description.ilike(search_pattern)) |
            (TestCase.preconditions.ilike(search_pattern)) |
            (TestCase.expected_result.ilike(search_pattern))
        )

    # 按创建时间倒序排序
    query = query.order_by(TestCase.created_at.desc())

    # 获取总数
    total = query.count()

    # 获取分页数据
    test_cases = query.offset(skip).limit(limit).all()
    
    return PaginatedResponse(
        items=test_cases,
        total=total,
        skip=skip,
        limit=limit
    )


@router.get("/{test_case_id}", response_model=TestCaseSchema)
def read_test_case(
    test_case_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """获取测试用例详情"""
    test_case = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        TestCase.id == test_case_id,
        Requirement.user_id == current_user.id
    ).first()
    
    if not test_case:
        raise HTTPException(status_code=404, detail="Test case not found")
    
    return test_case


@router.post("/", response_model=TestCaseSchema)
def create_test_case(
    test_case_in: TestCaseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """手动创建测试用例"""
    # 验证测试点是否属于当前用户
    test_point = db.query(TestPoint).join(Requirement).filter(
        TestPoint.id == test_case_in.test_point_id,
        Requirement.user_id == current_user.id
    ).first()

    if not test_point:
        raise HTTPException(status_code=404, detail="Test point not found")

    # 生成编号
    code = generate_test_case_code(db, test_point)

    test_case = TestCase(**test_case_in.model_dump(), code=code)
    db.add(test_case)
    db.commit()
    db.refresh(test_case)

    return test_case


@router.post("/generate/{test_point_id}")
async def generate_test_cases(
    test_point_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """为指定测试点生成测试用例"""
    test_point = db.query(TestPoint).join(Requirement).filter(
        TestPoint.id == test_point_id,
        Requirement.user_id == current_user.id
    ).first()
    
    if not test_point:
        raise HTTPException(status_code=404, detail="Test point not found")
    
    # 后台生成测试用例
    loop = asyncio.get_running_loop()
    background_tasks.add_task(
        generate_test_cases_background,
        test_point_id,
        current_user.id,
        loop
    )
    
    return {"message": "Generating test cases..."}


@router.put("/{test_case_id}", response_model=TestCaseSchema)
def update_test_case(
    test_case_id: int,
    test_case_in: TestCaseUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """更新测试用例"""
    test_case = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        TestCase.id == test_case_id,
        Requirement.user_id == current_user.id
    ).first()

    if not test_case:
        raise HTTPException(status_code=404, detail="Test case not found")

    update_data = test_case_in.model_dump(exclude_unset=True)
    # 确保不更新 code 字段（编号是自动生成的，不可修改）
    update_data.pop('code', None)

    for field, value in update_data.items():
        setattr(test_case, field, value)

    db.commit()
    db.refresh(test_case)

    return test_case


@router.delete("/{test_case_id}")
def delete_test_case(
    test_case_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """删除测试用例"""
    test_case = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        TestCase.id == test_case_id,
        Requirement.user_id == current_user.id
    ).first()
    
    if not test_case:
        raise HTTPException(status_code=404, detail="Test case not found")
    
    db.delete(test_case)
    db.commit()

    return {"message": "Test case deleted successfully"}


@router.post("/{test_case_id}/approve", response_model=TestCaseSchema)
def approve_test_case(
    test_case_id: int,
    approval: TestCaseApproval,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """审批测试用例"""
    # 验证审批状态
    if approval.approval_status not in ['approved', 'rejected']:
        raise HTTPException(status_code=400, detail="Invalid approval status. Must be 'approved' or 'rejected'")

    # 查询测试用例
    test_case = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        TestCase.id == test_case_id,
        Requirement.user_id == current_user.id
    ).first()

    if not test_case:
        raise HTTPException(status_code=404, detail="Test case not found")

    # 更新审批信息
    test_case.approval_status = approval.approval_status
    test_case.approved_by = current_user.id
    test_case.approved_at = func.now()
    test_case.approval_comment = approval.approval_comment

    db.commit()
    db.refresh(test_case)

    return test_case


@router.post("/{test_case_id}/reset-approval", response_model=TestCaseSchema)
def reset_test_case_approval(
    test_case_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """重置测试用例审批状态"""
    test_case = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        TestCase.id == test_case_id,
        Requirement.user_id == current_user.id
    ).first()

    if not test_case:
        raise HTTPException(status_code=404, detail="Test case not found")

    # 重置审批信息
    test_case.approval_status = 'pending'
    test_case.approved_by = None
    test_case.approved_at = None
    test_case.approval_comment = None

    db.commit()
    db.refresh(test_case)

    return test_case


@router.get("/export/excel")
def export_test_cases_to_excel(
    requirement_id: Optional[int] = None,
    test_point_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """导出测试用例到Excel"""
    print(f"[INFO] 开始导出测试用例 - requirement_id: {requirement_id}, test_point_id: {test_point_id}")

    # 构建查询
    query = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        Requirement.user_id == current_user.id
    )

    # 按需求筛选
    if requirement_id:
        query = query.filter(Requirement.id == requirement_id)

    # 按测试点筛选
    if test_point_id:
        query = query.filter(TestCase.test_point_id == test_point_id)

    test_cases = query.order_by(TestCase.code).all()

    if not test_cases:
        raise HTTPException(status_code=404, detail="No test cases found")

    print(f"[INFO] 找到 {len(test_cases)} 个测试用例")

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    cell_font = Font(name='微软雅黑', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置表头
    headers = [
        '需求名称', '测试点编号', '测试点名称', '测试用例编号', '测试用例标题',
        '描述', '前置条件', '测试步骤', '预期结果', '优先级', '测试类型',
        '审批状态', '审批意见', '创建时间'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置列宽
    column_widths = [20, 15, 25, 15, 30, 40, 30, 50, 30, 10, 12, 12, 30, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # 填充数据
    for row_num, test_case in enumerate(test_cases, 2):
        test_point = test_case.test_point
        requirement = test_point.requirement

        # 格式化测试步骤
        test_steps_text = ""
        if test_case.test_steps:
            for i, step in enumerate(test_case.test_steps, 1):
                action = step.get('action', '')
                expected = step.get('expected', '')
                test_steps_text += f"{i}. {action}"
                if expected:
                    test_steps_text += f"\n   预期: {expected}"
                test_steps_text += "\n"

        # 审批状态映射
        approval_status_map = {
            'pending': '待审批',
            'approved': '已通过',
            'rejected': '已拒绝'
        }

        # 优先级映射
        priority_map = {
            'high': '高',
            'medium': '中',
            'low': '低'
        }

        row_data = [
            requirement.title,
            test_point.code,
            test_point.title,
            test_case.code,
            test_case.title,
            test_case.description or '',
            test_case.preconditions or '',
            test_steps_text.strip(),
            test_case.expected_result or '',
            priority_map.get(test_case.priority, test_case.priority),
            test_case.test_type or '',
            approval_status_map.get(test_case.approval_status, test_case.approval_status),
            test_case.approval_comment or '',
            test_case.created_at.strftime('%Y-%m-%d %H:%M:%S') if test_case.created_at else ''
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 设置行高
    ws.row_dimensions[1].height = 30
    for row_num in range(2, len(test_cases) + 2):
        ws.row_dimensions[row_num].height = 60

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    from urllib.parse import quote

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if requirement_id:
        requirement = db.query(Requirement).filter(Requirement.id == requirement_id).first()
        filename = f"测试用例_{requirement.title}_{timestamp}.xlsx"
    elif test_point_id:
        test_point = db.query(TestPoint).filter(TestPoint.id == test_point_id).first()
        filename = f"测试用例_{test_point.title}_{timestamp}.xlsx"
    else:
        filename = f"测试用例_{timestamp}.xlsx"

    # URL编码文件名
    encoded_filename = quote(filename)

    print(f"[INFO] Excel 导出成功: {filename}")

    # 返回文件流
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.post("/{test_case_id}/match-scenario")
def match_scenario(
    test_case_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    为测试用例匹配场景
    
    根据测试用例的业务线、标题、描述等信息，智能匹配最合适的场景
    """
    # 查询测试用例及其关联信息
    test_case = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        TestCase.id == test_case_id,
        Requirement.user_id == current_user.id
    ).first()
    
    if not test_case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 获取测试点的业务线
    test_point = test_case.test_point
    business_line = test_point.business_line
    
    # 查询匹配的场景
    scenarios_query = db.query(Scenario).filter(Scenario.is_active == True)
    
    # 如果有业务线信息，优先匹配相同业务线的场景
    if business_line:
        scenarios_query = scenarios_query.filter(Scenario.business_line == business_line)
    
    scenarios = scenarios_query.all()
    
    if not scenarios:
        # 如果没有匹配的场景，返回所有启用的场景
        scenarios = db.query(Scenario).filter(Scenario.is_active == True).all()
    
    # 智能匹配：基于关键词匹配最相关的场景
    matched_scenario = None
    if scenarios:
        # 简单的关键词匹配逻辑
        test_case_keywords = set((test_case.title + ' ' + (test_case.description or '')).lower().split())
        
        max_score = 0
        for scenario in scenarios:
            scenario_keywords = set((scenario.name + ' ' + (scenario.description or '')).lower().split())
            # 计算关键词重叠度
            overlap = len(test_case_keywords & scenario_keywords)
            if overlap > max_score:
                max_score = overlap
                matched_scenario = scenario
        
        # 如果没有关键词重叠，返回第一个场景
        if not matched_scenario:
            matched_scenario = scenarios[0]
    
    if not matched_scenario:
        return {
            "matched": False,
            "message": "未找到匹配的场景，请先创建场景",
            "test_case": {
                "id": test_case.id,
                "code": test_case.code,
                "title": test_case.title,
                "business_line": business_line
            }
        }
    
    return {
        "matched": True,
        "scenario_code": matched_scenario.scenario_code,
        "scenario_name": matched_scenario.name,
        "scenario": {
            "id": matched_scenario.id,
            "scenario_code": matched_scenario.scenario_code,
            "name": matched_scenario.name,
            "description": matched_scenario.description,
            "business_line": matched_scenario.business_line,
            "channel": matched_scenario.channel,
            "module": matched_scenario.module
        },
        "test_case": {
            "id": test_case.id,
            "code": test_case.code,
            "title": test_case.title,
            "business_line": business_line
        },
        "message": f"成功匹配到场景: {matched_scenario.scenario_code}"
    }


@router.post("/{test_case_id}/generate-automation")
def generate_automation_case(
    test_case_id: int,
    module_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    生成自动化测试用例
    
    1. 匹配场景
    2. 调用自动化平台创建用例
    3. 获取支持的字段
    
    参数:
    - module_id: 模块ID，可选。如果不传，则从系统配置中读取
    """
    # 检查自动化服务是否可用
    if not automation_service:
        raise HTTPException(
            status_code=503,
            detail="自动化测试平台服务未配置或不可用"
        )
    
    # 如果没有传module_id，从系统配置读取
    if not module_id:
        module_id_config = db.query(SystemConfig).filter(
            SystemConfig.config_key == "AUTOMATION_PLATFORM_MODULE_ID"
        ).first()
        
        if module_id_config and module_id_config.config_value:
            module_id = module_id_config.config_value
        else:
            raise HTTPException(
                status_code=400,
                detail="模块ID未提供，且系统配置中未配置默认模块ID，请先在系统配置中配置自动化测试平台的模块ID"
            )
    
    # 查询测试用例及其关联信息
    test_case = db.query(TestCase).join(TestPoint).join(Requirement).filter(
        TestCase.id == test_case_id,
        Requirement.user_id == current_user.id
    ).first()
    
    if not test_case:
        raise HTTPException(status_code=404, detail="测试用例不存在")
    
    # 1. 匹配场景
    test_point = test_case.test_point
    business_line = test_point.business_line
    
    scenarios_query = db.query(Scenario).filter(Scenario.is_active == True)
    if business_line:
        scenarios_query = scenarios_query.filter(Scenario.business_line == business_line)
    
    scenarios = scenarios_query.all()
    
    if not scenarios:
        scenarios = db.query(Scenario).filter(Scenario.is_active == True).all()
    
    if not scenarios:
        raise HTTPException(
            status_code=404,
            detail="未找到可用的场景，请先在场景管理中创建场景"
        )
    
    # 简单匹配：选择第一个匹配的场景
    matched_scenario = scenarios[0]
    
    # 2. 调用自动化平台创建用例
    try:
        # 准备用例名称和描述
        case_name = test_case.title
        case_description = test_case.description or ""
        
        # 构建场景ID（使用场景的数据库ID或自定义映射）
        # 注意：这里假设场景的ID可以直接使用，实际可能需要映射
        scene_id = matched_scenario.scenario_code  # 使用场景编号作为场景ID
        
        # 准备测试用例信息供AI匹配
        test_case_info = {
            "title": test_case.title,
            "description": test_case.description or "",
            "preconditions": test_case.preconditions or "",
            "test_steps": str(test_case.test_steps) if test_case.test_steps else "",
            "expected_result": test_case.expected_result or "",
            "test_type": test_case.test_type or "",
            "priority": test_case.priority or ""
        }
        
        result = automation_service.create_case_with_fields(
            name=case_name,
            module_id=module_id,
            scene_id=scene_id,
            scenario_type="API",
            description=case_description,
            test_case_info=test_case_info
        )
        
        return {
            "success": True,
            "message": "AI智能匹配并成功创建自动化用例（含明细）",
            "data": {
                "test_case": {
                    "id": test_case.id,
                    "code": test_case.code,
                    "title": test_case.title
                },
                "matched_scenario": {
                    "id": matched_scenario.id,
                    "scenario_code": matched_scenario.scenario_code,
                    "name": matched_scenario.name
                },
                "selected_template": result.get("selected_template", {}),
                "created_case": result.get("created_case", {}),
                "template_case": result.get("template_case", {}),
                "case_detail": result.get("case_detail", {}),
                "supported_fields": result.get("fields", []),
                "new_usercase_id": result.get("new_usercase_id"),
                "template_usercase_id": result.get("template_usercase_id"),
                "scene_id": result.get("scene_id")
            }
        }
        
    except Exception as e:
        print(f"[ERROR] 创建自动化用例失败: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"创建自动化用例失败: {str(e)}"
        )

