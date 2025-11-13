from typing import List, Optional, Dict, Callable, Tuple
import zipfile
from xml.etree import ElementTree as ET

from docx import Document
from docx.oxml.ns import qn
import PyPDF2
import openpyxl
from langchain_community.document_loaders import (
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
    UnstructuredExcelLoader,
    UnstructuredFileLoader,
)


class DocumentParser:
    """文档解析服务，负责不同格式文档的文本提取"""

    MAX_CONSECUTIVE_EMPTY_ROWS = 2000
    WORD_NAMESPACE = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"

    @staticmethod
    def _collect_docx_text(doc: Document) -> List[str]:
        """Collect text from paragraphs, tables, headers, footers, and text boxes."""

        def append_text(target: List[str], text: str):
            text = text.strip()
            if text:
                target.append(text)

        def extract_from_table(table, target: List[str]):
            for row in table.rows:
                row_cells: List[str] = []
                seen_cells = set()
                for cell in row.cells:
                    cell_id = id(getattr(cell, "_tc", cell))
                    if cell_id in seen_cells:
                        continue
                    seen_cells.add(cell_id)
                    cell_text = " ".join(
                        p.text.strip() for p in cell.paragraphs if p.text.strip()
                    )
                    if cell_text:
                        row_cells.append(cell_text)
                    extract_from_container(cell, target)
                if row_cells:
                    target.append(" | ".join(row_cells))

        def extract_from_container(container, target: List[str]):
            if container is None:
                return
            for paragraph in getattr(container, "paragraphs", []):
                append_text(target, paragraph.text)
            for table in getattr(container, "tables", []):
                extract_from_table(table, target)

        blocks: List[str] = []
        extract_from_container(doc, blocks)

        txbx_tag = qn("w:txbxContent")
        paragraph_tag = qn("w:p")
        text_tag = qn("w:t")
        for txbx in doc.element.iter(txbx_tag):
            for para in txbx.iter(paragraph_tag):
                texts = [node.text for node in para.iter(text_tag) if node.text]
                merged = "".join(texts).strip()
                if merged:
                    blocks.append(merged)

        for section in getattr(doc, "sections", []):
            extract_from_container(getattr(section, "header", None), blocks)
            extract_from_container(getattr(section, "footer", None), blocks)
        return blocks

    @staticmethod
    def _load_with_langchain(loader) -> str:
        documents = loader.load()
        contents: List[str] = []
        for doc in documents:
            text = (getattr(doc, "page_content", "") or "").strip()
            if text:
                contents.append(text)
        return "\n".join(contents)

    @staticmethod
    def _run_attempts(attempts: List[Tuple[str, Callable[[], str]]], label: str) -> str:
        last_error: Optional[Exception] = None
        for source, func in attempts:
            try:
                result = func()
                if result and result.strip():
                    text = result.strip()
                    print(f"[INFO] {label} 解析成功（{source}），文本长度 {len(text)}")
                    return text
            except Exception as exc:
                last_error = exc
                print(f"[WARNING] {label} 解析失败（{source}）: {exc}")
        if last_error:
            raise last_error
        raise ValueError(f"{label} 未提取到任何文本内容")

    @classmethod
    def _extract_text_from_xml_bytes(cls, data: bytes) -> List[str]:
        try:
            root = ET.fromstring(data)
        except ET.ParseError:
            return []
        namespace = f"{{{cls.WORD_NAMESPACE}}}"
        paragraphs: List[str] = []
        for para in root.iter(f"{namespace}p"):
            texts = []
            for node in para.iter(f"{namespace}t"):
                if node.text:
                    texts.append(node.text)
            if texts:
                merged = "".join(texts).strip()
                if merged:
                    paragraphs.append(merged)
        return paragraphs

    @classmethod
    def _extract_docx_xml_text(cls, file_path: str) -> List[str]:
        xml_texts: List[str] = []
        try:
            with zipfile.ZipFile(file_path) as docx_zip:
                names = [
                    name
                    for name in docx_zip.namelist()
                    if name.startswith("word/")
                    and name.endswith(".xml")
                    and not name.endswith(".rels")
                ]
                priority = (
                    "word/document.xml",
                    "word/footnotes.xml",
                    "word/endnotes.xml",
                    "word/comments.xml",
                    "word/header1.xml",
                    "word/footer1.xml",
                )
                names.sort(
                    key=lambda n: (priority.index(n) if n in priority else len(priority), n)
                )
                for name in names:
                    try:
                        data = docx_zip.read(name)
                        xml_texts.extend(cls._extract_text_from_xml_bytes(data))
                    except KeyError:
                        continue
        except Exception as exc:
            print(f"[WARNING] DOCX XML 解析失败: {exc}")
        return xml_texts

    @classmethod
    def _parse_docx_native(cls, file_path: str) -> str:
        doc_blocks: List[str] = []
        try:
            doc = Document(file_path)
            doc_blocks = cls._collect_docx_text(doc)
        except Exception as exc:
            print(f"[WARNING] python-docx 解析异常: {exc}")

        xml_blocks = cls._extract_docx_xml_text(file_path)
        combined = doc_blocks + xml_blocks
        deduped: List[str] = []
        seen = set()
        for block in combined:
            stripped = block.strip()
            if not stripped:
                continue
            if stripped in seen:
                continue
            seen.add(stripped)
            deduped.append(stripped)

        if not deduped:
            raise ValueError("DOCX 文件未解析到任何文本内容")
        return "\n".join(deduped)

    @staticmethod
    def _parse_pdf_native(file_path: str) -> str:
        text: List[str] = []
        with open(file_path, "rb") as file_obj:
            pdf_reader = PyPDF2.PdfReader(file_obj)
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text.append(page_text)
        result = "\n".join(text).strip()
        if result:
            return result
        print("[WARNING] PyPDF2 未提取到有效内容，尝试 pdfminer.high_level.extract_text")
        from pdfminer.high_level import extract_text as pdfminer_extract_text

        fallback = pdfminer_extract_text(file_path).strip()
        if fallback:
            print("[INFO] pdfminer 解析成功，返回全文本")
            return fallback
        raise ValueError("PDF 文档未提取到任何文本")

    @staticmethod
    def _parse_txt_native(file_path: str) -> str:
        try:
            with open(file_path, "r", encoding="utf-8") as file_obj:
                return file_obj.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, "r", encoding="gbk") as file_obj:
                    return file_obj.read()
            except Exception:
                with open(file_path, "r", encoding="latin-1", errors="ignore") as file_obj:
                    return file_obj.read()

    @staticmethod
    def _normalize_cell(value) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    @classmethod
    def _build_excel_headers(cls, row) -> List[str]:
        headers: List[str] = []
        for idx, cell in enumerate(row):
            text = cls._normalize_cell(cell)
            if not text:
                text = f"列{idx + 1}"
            headers.append(text)
        return headers if headers else ["列1"]

    @classmethod
    def _format_excel_row(cls, row, headers: List[str]) -> Optional[str]:
        if not row:
            return None

        cells = []
        for idx, cell in enumerate(row):
            cell_text = cls._normalize_cell(cell)
            if not cell_text:
                continue
            header_name = headers[idx] if idx < len(headers) else f"列{idx + 1}"
            cells.append(f"{header_name}: {cell_text}")

        if not cells:
            return None
        return " | ".join(cells)

    @classmethod
    def _is_header_like_row(cls, headers: List[str], row) -> bool:
        total = 0
        matches = 0
        for idx, cell in enumerate(row):
            cell_text = cls._normalize_cell(cell)
            if not cell_text:
                continue
            total += 1
            header_name = headers[idx] if idx < len(headers) else f"列{idx + 1}"
            if cell_text == header_name:
                matches += 1
        return total > 0 and matches / total >= 0.8

    @classmethod
    def _parse_excel_native(cls, file_path: str) -> str:
        text: List[str] = []
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        print(f"[INFO] Excel 包含 {len(sheet_names)} 个 sheet: {', '.join(sheet_names)}")

        try:
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                text.append(f"Sheet: {sheet_name}")
                headers: List[str] = []
                header_row = None
                total_rows = 0
                effective_rows = 0
                empty_rows_after_data = 0

                for row in sheet.iter_rows(values_only=True):
                    total_rows += 1

                    if not headers:
                        if not any(cls._normalize_cell(cell) for cell in row):
                            continue
                        headers = cls._build_excel_headers(row)
                        header_row = row
                        text.append("Columns: " + " | ".join(headers))
                        row_text = cls._format_excel_row(row, headers)
                        if row_text and not cls._is_header_like_row(headers, row):
                            effective_rows += 1
                            text.append(row_text)
                        continue

                    row_text = cls._format_excel_row(row, headers)
                    if row_text:
                        effective_rows += 1
                        text.append(row_text)
                        empty_rows_after_data = 0
                    else:
                        empty_rows_after_data += 1
                        if empty_rows_after_data >= cls.MAX_CONSECUTIVE_EMPTY_ROWS:
                            print(
                                f"[INFO] 工作表 '{sheet_name}' 连续空行超过 {cls.MAX_CONSECUTIVE_EMPTY_ROWS}，提前结束解析"
                            )
                            break

                if effective_rows == 0 and header_row:
                    row_text = cls._format_excel_row(header_row, headers)
                    if row_text:
                        text.append(row_text)
                        effective_rows = 1

                print(
                    f"[INFO] 工作表 '{sheet_name}' 解析完成，提取 {effective_rows} 行（总{total_rows}行）"
                )
        finally:
            workbook.close()

        return "\n".join(text)

    @classmethod
    def parse_docx(cls, file_path: str) -> str:
        attempts: List[Tuple[str, Callable[[], str]]] = [
            ("langchain-docx2txt", lambda: cls._load_with_langchain(Docx2txtLoader(file_path))),
            ("langchain-unstructured", lambda: cls._load_with_langchain(UnstructuredFileLoader(file_path, mode="elements"))),
            ("native", lambda: cls._parse_docx_native(file_path)),
        ]
        return cls._run_attempts(attempts, "DOCX")

    @classmethod
    def parse_pdf(cls, file_path: str) -> str:
        attempts = [
            ("langchain-pypdf", lambda: cls._load_with_langchain(PyPDFLoader(file_path))),
            ("langchain-unstructured", lambda: cls._load_with_langchain(UnstructuredFileLoader(file_path, mode="elements"))),
            ("native", lambda: cls._parse_pdf_native(file_path)),
        ]
        return cls._run_attempts(attempts, "PDF")

    @classmethod
    def parse_txt(cls, file_path: str) -> str:
        attempts = [
            ("langchain-text", lambda: cls._load_with_langchain(TextLoader(file_path, autodetect_encoding=True))),
            ("native", lambda: cls._parse_txt_native(file_path)),
        ]
        return cls._run_attempts(attempts, "TXT")

    @classmethod
    def parse_excel(cls, file_path: str) -> str:
        attempts = [
            ("langchain-unstructured-excel", lambda: cls._load_with_langchain(UnstructuredExcelLoader(file_path))),
            ("native", lambda: cls._parse_excel_native(file_path)),
        ]
        return cls._run_attempts(attempts, "EXCEL")

    @classmethod
    def parse(cls, file_path: str, file_type: str) -> Optional[str]:
        parsers = {
            "docx": cls.parse_docx,
            "pdf": cls.parse_pdf,
            "txt": cls.parse_txt,
            "xls": cls.parse_excel,
            "xlsx": cls.parse_excel,
        }

        parser = parsers.get(file_type.lower())
        if parser:
            return parser(file_path)
        return None

    @staticmethod
    def evaluate_quality(text: str) -> Dict[str, float]:
        """返回文本质量指标，用于校验"""
        lines = text.splitlines()
        total_lines = len(lines)
        non_empty_lines = sum(1 for line in lines if line.strip())
        non_empty_ratio = (non_empty_lines / total_lines) if total_lines else 0.0
        meaningful_chars = sum(1 for ch in text if not ch.isspace())
        return {
            "total_lines": total_lines,
            "non_empty_lines": non_empty_lines,
            "non_empty_ratio": non_empty_ratio,
            "meaningful_chars": meaningful_chars,
        }
